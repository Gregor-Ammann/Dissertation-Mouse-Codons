from Bio import SeqIO
import openpyxl


class Proteins:
    def __init__(self, seq_record):

        # Initialize a dictionary, in which codons are keys and the count of this codon in the sequence are values.
        self.all_codons_in_sequence = {}

        # The .fasta-file contains DNA data, which can be transcribed into RNA. (T is exchanged for U)
        transcribed_sequence = seq_record.seq.transcribe()

        # First, the sequence is split into codons. The number of each codon is counted and stored into the
        # dictionary named "all_codons_in_sequence".
        for i in range(0, len(seq_record.seq), 3):
            current_codon = str(transcribed_sequence[i:i + 3])
            self.all_codons_in_sequence.setdefault(current_codon, 0)
            self.all_codons_in_sequence[current_codon] += 1

        # To make sure that AAA, CAA, GAA, AAG, CAG and GAG are part of the dictionary
        important_codons = ["AAA", "CAA", "GAA", "AAG", "CAG", "GAG"]
        for i in range(len(important_codons)):
            self.all_codons_in_sequence.setdefault(important_codons[i], 0)

        # int sum_of_affected_codons
        self.sum_of_affected_codons = self.all_codons_in_sequence["AAA"] + self.all_codons_in_sequence["CAA"] \
                                      + self.all_codons_in_sequence["GAA"]

        # int sum_of_unaffected codons
        self.sum_of_unaffected_codons = self.all_codons_in_sequence["AAG"] + self.all_codons_in_sequence["CAG"] \
                                        + self.all_codons_in_sequence["GAG"]

        # int ratio affected / (affected + unaffected)
        if self.sum_of_affected_codons == 0:
            self.ratio_affected = 0
        else:
            self.ratio_affected = self.sum_of_affected_codons / (self.sum_of_affected_codons
                                                                 + self.sum_of_unaffected_codons)

        # For RSCU analysis, I need a list that contains ONLY the percentage of affected codons within each Transcript
        # This percentage is then divided by 50% to calculate RSCU of each protein
        self.RSCU = self.ratio_affected / 0.5

        # int ratio affected/length
        self.affected_per_length = self.sum_of_affected_codons / len(str(seq_record))

        # int ratio unaffected/length
        self.unaffected_per_length = self.sum_of_unaffected_codons / len(str(seq_record))

        # CCDS numbers are retrieved from sequence.id
        self.id = str(seq_record.id)[:11]

"""
# Calculates codon bias based on number of A-ending codons ("affected")
def calculate_codon_bias_absolute(all_proteins_and_codons, highest_absolute):
    all_proteins_and_codons.sort(key=lambda x: x.sum_of_affected_codons, reverse=True)
    for i in range(top_x):
        highest_absolute.append(all_proteins_and_codons[i])


# Calculates codon bias based on number of A-ending codons ("affected") divided by (A-ending + G-ending) codons
def calculate_codon_bias_relative(all_proteins_and_codons, highest_relative):
    all_proteins_and_codons.sort(key=lambda x: x.ratio_affected, reverse=True)
    for i in range(top_x):
        highest_relative.append(all_proteins_and_codons[i])


# Calculates codon bias based on number of A-ending codons ("affected") divided by mRNA length
def calculate_codon_bias_length(all_proteins_and_codons, highest_length):
    all_proteins_and_codons.sort(key=lambda x: x.unaffected_per_length, reverse=True)
    for i in range(top_x):
        highest_length.append(all_proteins_and_codons[i])
"""

# Calculate RSCU based on A-ending codons ("affected") divided by A and G ending codons
def calculate_RSCU(all_proteins_and_codons, RSCU_list):
    for i in range(len(all_proteins_and_codons)):
        RSCU_list.append(all_proteins_and_codons[i].RSCU)


# Initialize some variables needed for calculation of codon bias in different ways
all_proteins_and_codons = []
highest_absolute = []
highest_relative = []
highest_length = []
RSCU_list = []

# Number of genes with the most codon bias
top_x = 100

# Open and read the .fasta file. Store the data as Protein-objects in the list called all_proteins_and_codons
for seq_record in SeqIO.parse("CCDS_nucleotide.current.fna", "fasta"):
    protein = Proteins(seq_record)
    all_proteins_and_codons.append(protein)
# print(len(all_proteins_and_codons))

# calculate the codon bias
#calculate_codon_bias_length(all_proteins_and_codons, highest_length)
# calculate_codon_bias_relative(all_proteins_and_codons, highest_relative)

# Calculate RSCU
calculate_RSCU(all_proteins_and_codons, RSCU_list)

# int list with CCDS names of the most affected proteins
highest_length_CCDS_name = []
highest_relative_CCDS_name = []

"""# append the ID of all Proteins-class elements in the top_x list to this new list
for i in range(len(highest_length)):
    highest_relative_CCDS_name.append(highest_length[i].id)"""

# Open Excel sheet to retrieve gene names
wb = openpyxl.load_workbook("CCDS.current..xlsx")
sheet = wb["CCDS.current"]

# Check for each gene, if it is in the codon bias list. If it is, get the gene name (not CCDS name) and save it.
# Note, that this list is not sorted by codon bias.
gene_names = []
for i in range(1, sheet.max_row + 1):
    ccds_name = sheet.cell(row=i, column=5).value
    if ccds_name in highest_relative_CCDS_name:
        gene_name = sheet.cell(row=i, column=3).value
        gene_names.append(gene_name)

# Open input worksheet and select current sheet
input_genes = openpyxl.load_workbook("cortex.xlsx")
input_sheet = input_genes.active

# Iterate through all gene names and append to list_of_input_genes
list_of_input_genes = []
for i in range(1, input_sheet.max_row + 1):
    try:
        list_of_input_genes.append(input_sheet.cell(row=i, column=1).value.lower())
    except:
        pass

# Create a dictionary that translates gene names into CCDS names with gene as key and CCDS as value
# Make gene names lowercase
gene_CCDS_translator = {}
for i in range(1, sheet.max_row + 1):
    gene_CCDS_translator[sheet.cell(row=i, column=3).value.lower()] = sheet.cell(row=i, column=5).value.lower()

# Liver CCDS list
liver_CCDS = []
# Check, if the genes in liver.xlsx are in the dictionary. If they are, add the respective CCDS to the list
for input_gene in list_of_input_genes:
    if input_gene in gene_CCDS_translator.keys():
        liver_CCDS.append(gene_CCDS_translator[input_gene])

# RSCU value list
liver_RSCU = []
# Calculate the RSCU for each protein found upregulated in liver
for protein in all_proteins_and_codons:
    if protein.id.lower() in liver_CCDS:
        liver_RSCU.append(protein.RSCU)

# Safe RSCU into a excel file
wb_liver = openpyxl.Workbook()
sheet_liver = wb_liver.active
for i in range(len(liver_RSCU)):
    sheet_liver.cell(row=i+1, column=1).value = liver_RSCU[i]
sheet_liver.cell(row=1, column=2).value = "Total Proteins: " + str(len(liver_RSCU))
wb_liver.save("cortex.xlsx")

total_proteome = openpyxl.Workbook()
total_sheet = total_proteome.active
for i in range(len(all_proteins_and_codons)):
    total_sheet.cell(row=i+1, column=1).value = all_proteins_and_codons[i].RSCU
total_proteome.save("totalporteome_results.xlsx")




# Save the top_x genes in a new Excelsheet.
# Create Excel document
wb_results = openpyxl.Workbook()
sheet_results = wb_results.active

# Add results into each row of column 1
for i in range(len(gene_names)):
    sheet_results.cell(row=i + 1, column=1).value = gene_names[i]
# Save the Excelsheet
wb_results.save("test_results.xlsx")
